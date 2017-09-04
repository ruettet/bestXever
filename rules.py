from openpyxl import load_workbook
from re import sub
from codecs import open


class BestClassEver(object):

    def __init__(self):
        self.punctuation = (".", ",", ";", ":", "!", "?", "-", "=")
        self.junction = ("en", "maar", "and", "but", "doch")
        self.reference = ("die", "dat", "wat", "wie", "wiens", "hetgeen", "hetgene")
        self.auxiliaries = ("ben", "bent", "is", "zijn", "was", "waren", "heb", "hebt", "hebben", "had", "hadden",
                            "word", "wordt", "worden", "werd", "werden", "kan", "kunt", "kunnen", "kon", "konden",
                            "moest", "moesten", "mag", "mogen", "mocht", "mochten", "zal", "zullen","zou", "zouden")
        self.pronouns = ("ik", "jij", "je", "we", "wij", "zij")
        self.alg_high_freq = ("er", "van", "in", "op", "met", "voor", "aan", "dan")
        self.highfreq_os = ("want", "omdat", "terwijl", "als", "zoals")
        self.headers = ["random", "left_context", "targetconstructie", "right_context", "rule_1_1", "rule_1_2",
                        "rule_1_3", "rule_1_4", "rule_2_1", "rule_2_2", "rule_2_3", "rule_2_4", "rule_2_5", "rule_2_6"]

    def load_data(self):
        wb = load_workbook(filename='ez annotated tweets for paper.xlsx', read_only=True)
        ws = wb['ez splice OUT 3']
        i = 1
        amount_observations = len(list(ws.rows))
        for row in range(2, amount_observations+1):
            print(i, "of", amount_observations)
            i += 1
            random = ws.cell(row=row, column=1).value
            tekst = str(ws.cell(row=row, column=10).value)
            targetconstructie = [str(ws.cell(row=row, column=3).value),
                                 str(ws.cell(row=row, column=4).value),
                                 str(ws.cell(row=row, column=5).value),
                                 str(ws.cell(row=row, column=6).value)] \
                if ws.cell(row=row, column=2).value == 'withdet' \
                else [str(ws.cell(row=row, column=4).value),
                      str(ws.cell(row=row, column=5).value),
                      str(ws.cell(row=row, column=6).value)]
            targetconstructie_str =(" ".join(targetconstructie)).strip()
            try:
                start_index = tekst.index(targetconstructie_str)
                end_index = start_index + len(targetconstructie_str)
                left_context = self._prune_left_context(tekst[:start_index].strip()) if start_index > 0 else ''
                right_context = tekst[end_index:].strip() if end_index < len(tekst) else ''
                data = {
                    "random": random,
                    "targetconstructie": targetconstructie_str.strip(),
                    "left_context": left_context,
                    "right_context": right_context
                }
                data["rule_1_1"] = self.rule_1_1(data)
                data["rule_1_2"] = self.rule_1_2(data)
                data["rule_1_3"] = self.rule_1_3(data)
                data["rule_1_4"] = self.rule_1_4(data)
                data["rule_2_1"] = self.rule_2_1(data)
                data["rule_2_2"] = self.rule_2_2(data)
                data["rule_2_3"] = self.rule_2_3(data)
                data["rule_2_4"] = self.rule_2_4(data)
                data["rule_2_5"] = self.rule_2_5(data)
                data["rule_2_6"] = self.rule_2_6(data)

                with open("data.tsv", "a", "utf-8") as f:
                    line = []
                    for header in self.headers:
                        line.append(str(data[header]))
                    f.write("\t".join(line) + "\n")

            except ValueError:
                print("could not check", data["random"])
                continue

    def rule_1_1(self, data):
        return data["left_context"].endswith(self.punctuation) and data["right_context"].startswith(self.punctuation)

    def rule_1_2(self, data):
        return (data["left_context"].endswith(self.punctuation) or data["left_context"].endswith(self.junction)) and \
               (data["right_context"].endswith(self.punctuation) or data["right_context"].startswith(self.junction))

    def rule_1_3(self, data):
        return data["left_context"] == "" and (data["right_context"] == "" or data["right_context"].startswith(self.punctuation))

    @staticmethod
    def rule_1_4(data):
        return data["left_context"].endswith("(") and data["right_context"].startswith(")")

    def rule_2_1(self, data):
        return data["left_context"].endswith(self.reference) or data["right_context"].startswith(self.reference)

    def rule_2_2(self, data):
        return data["left_context"].endswith(self.auxiliaries) or data["right_context"].startswith(self.auxiliaries)

    def rule_2_3(self, data):
        return data["left_context"].endswith(self.pronouns) or data["right_context"].startswith(self.pronouns)

    def rule_2_4(self, data):
        return data["left_context"].endswith(self.alg_high_freq) or data["right_context"].startswith(self.alg_high_freq)

    def rule_2_5(self, data):
        left_words = data["left_context"].split(" ")
        right_words = data["right_context"].split(" ")
        words = left_words[-3:] + right_words[:3]
        return any(highfreq_os in "".join(words) for highfreq_os in self.highfreq_os)

    @staticmethod
    def rule_2_6(data):
        left_words = sub(r"[^\w\s]", "", data["left_context"]).split(" ")
        right_words = sub(r"[^\w\s]", "", data["right_context"]).split(" ")
        return len(left_words) >= 3 or len(right_words) >= 3

    @staticmethod
    def _prune_left_context(left_context):
        allowed_left_contexts = ["een van", "één van", "1 van", "1", "echt", "aller"]
        for allowed_left_context in allowed_left_contexts:
            if left_context.endswith(allowed_left_context):
                return left_context[:len(left_context)-len(allowed_left_context)].strip()
        return left_context.strip()

if __name__ == "__main__":
    bce = BestClassEver()
    bce.load_data()
